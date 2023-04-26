# -*- coding: utf-8 -*-
"""
Created on Tue Feb 21 13:43:25 2023

@author: cail
"""

import requests
import os
import time
from PIL import Image
import pandas as pd
import openpyxl
from openpyxl import load_workbook
import tkinter as tk
from tkinter import filedialog
from tkinter import messagebox


df = pd.read_excel('PrefixeSocoda (003).xlsx')
supplier_to_brands = {}
for supplier, brand in zip(df['FABRICANT'], df['MARQUE']):
    if supplier not in supplier_to_brands:
        supplier_to_brands[supplier] = []
    if brand not in supplier_to_brands[supplier]:
        supplier_to_brands[supplier].append(brand)

# Mettre à jour la liste des marques en fonction du fournisseur sélectionné
def update_brand_list(*args):
    selected_supplier = supplier_var.get()
    brand_list['menu'].delete(0, 'end')
    brands = df[df['FABRICANT'] == selected_supplier]['MARQUE'].unique()
    for brand in brands:
        brand_list['menu'].add_command(label=brand, command=tk._setit(brand_var, brand))

def get_trigram(fabricant, marque):
    trigram = df.loc[(df['FABRICANT'] == fabricant) & (df['MARQUE'] == marque), 'PREFIXE'].values
    if len(trigram) == 0:
        return "Trigramme non trouvé"
    else:
        return trigram[0]
    
def show_error_popup(message):
    root = tk.Tk()
    root.withdraw()
    messagebox.showerror("Error", message)

def printProgressBar (iteration, total, prefix = '', suffix = '', decimals = 1, length = 100, fill = '█', printEnd = "\r"):

    percent = ("{0:." + str(decimals) + "f}").format(100 * (iteration / float(total)))
    filledLength = int(length * iteration // total)
    bar = fill * filledLength + '-' * (length - filledLength)

    print(f'\r{prefix} |{bar}| {percent}% {suffix}', end = printEnd)
    # Print New Line on Complete
    if iteration == total:
        print()

def xnviewConversion(chemin_dossier):
    file_list = os.listdir(chemin_dossier)
    maxim = len(file_list)
    if maxim == 0:
        show_error_popup("Aucune images n'a été trouvé.")
        raise ValueError("Aucune images n'a été trouvé.")
        
    tracer = 1
    printProgressBar(0, maxim, prefix = 'Progress:', suffix = 'Complete', length = 50)
    for file in file_list:
        tracer = tracer +1
        printProgressBar(tracer, maxim, prefix = 'Progress:', suffix = 'Complete', length = 50)
        if file.endswith(".jpg") or file.endswith(".jpeg") or file.endswith(".png"):
            file_path = os.path.join(chemin_dossier, file)
            # Ouvrir l'image
            image = Image.open(file_path)

            # Redimensionner l'image en 600x600 en conservant l'aspect ratio
            new_size = (600, 600)
            image.thumbnail(new_size)
            
            # Créer une nouvelle image de taille 600x600 avec un fond noir
            new_image = Image.new("RGB", new_size, color="white")
            
            # Coller l'image redimensionnée au centre de la nouvelle image avec un fond noir
            left = (new_size[0] - image.size[0]) // 2
            top = (new_size[1] - image.size[1]) // 2
            new_image.paste(image, (left, top))
            
            # Sauvegarder la nouvelle image
            new_image.save(file_path)
            
def mediator(wb, four_name, marq_name, trigramme):
    start_time = time.perf_counter()
    workbook = load_workbook(wb)
    sheet = workbook['03_MEDIA']
    z = 1 
    photobd = ['photobd', 'PHOTOBD']
    photohd = ['photohd', 'PHOTOHD']
    photohda = ['photohda', 'PHOTOHDA']
    photonorm = ['photo', 'PHOTO']
    ext2 = '.pdf'
    ext = '.jpg'
    colonne = recuperer_ltre('URLT', sheet['A:Z'])
    colonne2 = recuperer_ltre('TYPM', sheet['A:Z'])
    colonne3 = recuperer_ltre('NUM', sheet['A:Z'])
    colonne4 = recuperer_ltre('REFCIALE', sheet['A:Z'])
    colonne5 = recuperer_ltre('NOM', sheet['A:Z'])
    colonne6 = recuperer_ltre('TYPM', sheet['A:Z'])
    colonne7 = recuperer_ltre('MARQUE', sheet['A:Z'])
    dossier_parent  = os.path.dirname(wb)
    photo_folder = os.path.join(dossier_parent , 'Photo ' + str(marq_name))
    fiche_folder = os.path.join(dossier_parent , 'Fiche ' + str(marq_name))
    os.makedirs(photo_folder, exist_ok=True)
    os.makedirs(fiche_folder, exist_ok=True)
    photoaprendre = ''
    status = False
    error_df = pd.DataFrame(columns=["REFCIALE",'URL', 'TYPE'])
    error_file = os.path.join(dossier_parent, "error_"+str(trigramme)+".xlsx")
    workbook = openpyxl.Workbook()
    workbook.save(error_file)
    
    for row in sheet[colonne2]:
        if row.value in photobd:
            photoaprendre = row.value
            status = True
            break
        
    if status == False: 
        for row in sheet[colonne2]:
            if row.value in photohd:
                photoaprendre = row.value
                status = True
                break
    
    if status == False: 
        for row in sheet[colonne2]:
            if row.value in photohda:
                photoaprendre = row.value
                status = True
                break
            
    if status == False: 
        for row in sheet[colonne6]:
            if row.value in photonorm:
                photoaprendre = row.value
                status = True
                break
   
    max_row = sheet.max_row
    printProgressBar(0, max_row, prefix = 'Progress:', suffix = 'Complete', length = 50)
    
    # PHOTO  
    for row in sheet[colonne]:
        tracer = row.row
        printProgressBar(tracer, max_row, prefix = 'Progress:', suffix = 'Complete', length = 50)
        if row.value != 'URLT':
            if sheet[colonne7 + str(tracer)].value.lower() == marq_name.lower():
                if sheet[colonne2 + str(tracer)].value == photoaprendre:
                    if str(sheet[colonne3 + str(tracer)].value) == '1':
                        try:
                            sheet[colonne5 + str(tracer)].value[-4:] != None
                        except Exception :
                            nt = 1
                        ref = sheet[colonne4 + str(tracer)].value
                        photo_new_name = trigramme +'_'+ str(ref) + ext
                        photo_new_name = photo_new_name.replace("\\", "")
                        photo_new_name = photo_new_name.replace("/", "")
                        destfile = photo_folder
                        filedestination = os.path.join(destfile, photo_new_name)
                        url = row.value
                        if not os.path.exists(filedestination):
                            if url != None:
                                try:
                                    response = requests.get(url)
                                except Exception :
                                    error_df = error_df.append({'REFCIALE': ref, 'URL': url, 'TYPE': 'PHOTO'}, ignore_index=True)
                                else:   
                                    if response.status_code == 200:
                                        with open(filedestination, 'wb') as f:
                                            f.write(response.content)
                                            z = z +1
                                    else:
                                        error_df = error_df.append({'REFCIALE': ref, 'URL': url, 'TYPE': 'PHOTO'}, ignore_index=True)
                            if url == None:
                                z = z +1
    xnviewConversion(photo_folder)
    # FICHE 
    
    max_row = sheet.max_row
    printProgressBar(0, max_row, prefix = 'Progress:', suffix = 'Complete', length = 50)
    for row in sheet[colonne]:
        tracer = row.row
        printProgressBar(tracer, max_row, prefix = 'Progress:', suffix = 'Complete', length = 50)
        if row.value != 'URLT':
            if sheet[colonne7 + str(tracer)].value.lower() == marq_name.lower():
                if sheet[colonne2 + str(tracer)].value.lower() == 'fiche':
                    if str(sheet[colonne3 + str(tracer)].value) == '1':
                        try:
                            sheet[colonne5 + str(tracer)].value[-4:] != None
                        except Exception :
                            nt = 1
                        ref = sheet[colonne4 + str(tracer)].value
                        fiche_new_name = trigramme +'_'+ str(ref) + ext2
                        fiche_new_name = fiche_new_name.replace("\\", "_")
                        fiche_new_name = fiche_new_name.replace("/", "_")
                        destfile = fiche_folder
                        filedestination = os.path.join(destfile, fiche_new_name)
                        url = row.value
                        if not os.path.exists(filedestination):
                            if url != None:
                                try:
                                    response = requests.get(url)
                                except Exception :
                                    error_df = error_df.append({'REFCIALE': ref, 'URL': url, 'TYPE': 'FICHE'}, ignore_index=True)
                                else: 
                                    if response.status_code == 200:
                                        with open(filedestination, 'wb') as f:
                                            f.write(response.content)
                                            z = z +1
                                    else:
                                        error_df = error_df.append({'REFCIALE': ref, 'URL': url, 'TYPE': 'FICHE'}, ignore_index=True)
                            if url == None:
                                z = z +1
    if not error_df.empty:
        error_df.to_excel(error_file, index=False)
        print(f'Les erreurs ont été enregistrées dans {error_file}')
    else:
        print('Toutes les URLs ont été téléchargées avec succès !')
    end_time = time.perf_counter()
    execution_time = end_time - start_time
    print("----------------------------------------------------------")
    print("Temps d'exécution : {:.2f} seconds".format(execution_time))
    print("----------------------------------------------------------") 
    root.destroy()
    
def recuperer_ltre(valeur, plage):
    for row in plage:
        for cell in row:
            if cell.value == valeur :
                column = cell.column
                ltr_column = openpyxl.utils.get_column_letter(column)
                return ltr_column
        
def open_file():
    folder_path = filedialog.askopenfilename()
    folder_entry.delete(0, tk.END)
    folder_entry.insert(0, folder_path)

def update_progressbar(pbar, value):
    pbar['value'] = value
    pbar.update()

def clear_entry():
    folder_entry.delete(0, tk.END)
    valider_button.config(state='normal')
    folder_entry.config(state='normal')
    folder_button.config(state='normal')
    
def new_tarif():
    folder_entry.delete(0, tk.END)
    valider_button.config(state='normal')
    folder_entry.config(state='normal')
    folder_button.config(state='normal')
    clear_button.config(state='normal')

def OnValidation():
    folder_path = folder_entry.get()
    fourn_name = supplier_var.get()
    marq_name = brand_var.get()
    if folder_path != "":
        print("Chemin du dossier : "+folder_path)       
        if str(fourn_name) != "" or str(fourn_name) != "-Choisir un fournisseur-":
            print("Nom du fournisseur : "+fourn_name)
            if str(marq_name) != "" or str(marq_name) !="-Choisir une marque-":
                print("Nom de la marque : "+marq_name)
            else:
                print("Le nom de la marque n'est pas spécifié")
        else:
            print("Le nom du fournisseur n'est pas spécifié")
    else:
        print("Chemin du dossier n'est pas spécifié")
    
    if folder_path != "" and fourn_name != "" and marq_name != "":
        valider_button.config(state='disabled')
        folder_entry.config(state='disabled')
        folder_button.config(state='disabled')
        clear_button.config(state = "disabled")
        
        trigramme = get_trigram(fourn_name, marq_name).lower()
        mediator(folder_path, fourn_name, marq_name, trigramme)
         
def quit_app():
    root.destroy()
    
root = tk.Tk()
root.title("MEDIATOR 1.0.1")
root.geometry("450x160")

texte_param = tk.Label(root, text="Initialisation de Mediator")
texte_param.grid(row=0, column=1)

folder_label = tk.Label(root, text="Fichier FAB-DIS :")
folder_label.grid(row=1, column=0)

folder_entry = tk.Entry(root, width=30)
folder_entry.grid(row=1, column=1)

folder_button = tk.Button(root, text="   ...   ", command=open_file)
folder_button.grid(row=1, column=2)

supplier_label = tk.Label(root, text="Fournisseur")
supplier_label.grid(row=3, column=0)

supplier_var = tk.StringVar(root)
supplier_var.set("-Choisir un fournisseur-")

supplier_list = tk.OptionMenu(root, supplier_var, *df['FABRICANT'].unique())
supplier_list.grid(row=3, column=1)

brand_label = tk.Label(root, text="Marque")
brand_label.grid(row=4, column=0)

brand_var = tk.StringVar(root)
brand_var.set("-Choisir une marque-")

brand_list = tk.OptionMenu(root, brand_var, "")
brand_list.grid(row=4, column=1)

clear_button = tk.Button(root, text="Effacer", command=clear_entry, state= 'normal')
clear_button.grid(row=6, column=1)

valider_button = tk.Button(root, text="Valider",state='normal', command=OnValidation)
valider_button.grid(row=6, column=2)

quit_button = tk.Button(root, text="Quitter", command=quit_app)
quit_button.grid(row=7, column=2)

new_button = tk.Button(root, text="Nouveau ! ", command=new_tarif)
new_button.grid(row=7, column=0)

supplier_var.trace('w', update_brand_list)

root.mainloop()



