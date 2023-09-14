# -*- coding: utf-8 -*-
"""
Created on Tue Jul 11 15:18:40 2023

@author: cail
"""

import time
import pandas as pd
import openpyxl
from openpyxl import load_workbook
import tkinter as tk
from tkinter import filedialog
from tkinter import messagebox

df = pd.read_excel('PrefixeSocoda (003).xlsx')
supplier_to_brands = {}
for supplier, brand in zip(df['FABRICANT'], df['MARQUE']):
    if supplier not in supplier_to_brands:
        supplier_to_brands[supplier] = []
    if brand not in supplier_to_brands[supplier]:
        supplier_to_brands[supplier].append(brand)

# Mettre à jour la liste des marques en fonction du fournisseur sélectionné
def update_brand_list(*args):
    selected_supplier = supplier_var.get()
    brand_list['menu'].delete(0, 'end')
    brands = df[df['FABRICANT'] == selected_supplier]['MARQUE'].unique()
    for brand in brands:
        brand_list['menu'].add_command(label=brand, command=tk._setit(brand_var, brand))

def get_trigram(fabricant, marque):
    trigram = df.loc[(df['FABRICANT'] == fabricant) & (df['MARQUE'] == marque), 'PREFIXE'].values
    if len(trigram) == 0:
        return "Trigramme non trouvé"
    else:
        return trigram[0]

def get_code_fou(fabricant, marque):
    code_fou = df.loc[(df['FABRICANT'] == fabricant) & (df['MARQUE'] == marque), 'CODEFOU'].values
    if len(code_fou) == 0:
        return "Code fournisseur non trouvé"
    else:
        return code_fou
    
def show_error_popup(message):
    root = tk.Tk()
    root.withdraw()
    messagebox.showerror("Error", message)

def printProgressBar (iteration, total, prefix = '', suffix = '', decimals = 1, length = 100, fill = '█', printEnd = "\r"):

    percent = ("{0:." + str(decimals) + "f}").format(100 * (iteration / float(total)))
    filledLength = int(length * iteration // total)
    bar = fill * filledLength + '-' * (length - filledLength)

    print(f'\r{prefix} |{bar}| {percent}% {suffix}', end = printEnd)
    # Print New Line on Complete
    if iteration == total:
        print()


def create_request(wb, four_name, marq_name, trigramme):
    start_time = time.perf_counter()
    workbook = load_workbook(wb)
    sheet = workbook['01_COMMERCE']
    sheet2 = workbook.create_sheet(title="requete")
    sheet2["A1"] = "REQ DESC"
    sheet2["B1"] = "REQ PHOTO"
    code_fou = str(get_code_fou(four_name, marq_name))
    code_fou = code_fou.replace("[", "")
    code_fou = code_fou.replace("]", "")
    if len(code_fou) < 5 :
        code_fou = "0" + str(code_fou)
        print(code_fou)

    refciale = ""
    gtin13 = ""
    fab = four_name
    marq = marq_name
    gamme =""
    libelle30 = ""
    libelle240 = ""    

    colonne7 = recuperer_ltre('REQ DESC', sheet2['A:Z'])
    colonne9 = recuperer_ltre('REQ PHOTO', sheet2['A:Z'])
    colonne3 = recuperer_ltre('LIBELLE30', sheet['A:Z'])
    colonne4 = recuperer_ltre('LIBELLE240', sheet['A:Z'])
    colonne12 = recuperer_ltre('LIBELLE80', sheet['A:Z'])
    colonne5 = recuperer_ltre('GAMME', sheet['A:Z'])
    colonne6 = recuperer_ltre('REFCIALE', sheet['A:Z'])
    colonne8 = recuperer_ltre('GTIN13', sheet['A:Z'])
    colonne10 = recuperer_ltre('SOCODA', sheet['A:Z'])
    colonne11 = recuperer_ltre('PHOTO', sheet['A:Z'])
    
    
    for row in sheet[colonne6]:
        tracer = row.row
        refciale = sheet[colonne6 + str(tracer)].value
        libelle30 = sheet[colonne3 + str(tracer)].value
        socoda = sheet[colonne10 + str(tracer)].value
        photo = sheet[colonne11 + str(tracer)].value
        libelle240 = sheet[colonne4 + str(tracer)].value
        
        if libelle240 == "":
            libelle240 = sheet[colonne12 + str(tracer)].value
        
        libelle240 = libelle240.replace(";","</li><li>")
        libelle240 = libelle240.replace(".","</li><li>")
        libelle240 = libelle240.replace("+","</li><li>")
        libelle240 = libelle240.replace("-","</li><li>")
        libelle240 = libelle240.replace(",","</li><li>")
        libelle240 = libelle240.replace("'","")
        libelle30 = libelle30.replace("'"," ")

        gamme = sheet[colonne5 + str(tracer)].value
        gtin13 = sheet[colonne8 + str(tracer)].value
        htmldesc = "<p><b> Code fournisseur : "+str(code_fou)+"</b><br> Reférence commerciale : "+str(refciale)+"<br> Fournisseur : "+str(fab)+"<br>Nom produit : "+str(libelle30)+"<br>Marque : "+str(marq)+"<br>Gamme : "+str(gamme)+"<br>EAN : "+str(gtin13)+"</p><p><b>Caractéristiques :</b><br/><ul><li>"+str(libelle240)+"</li></ul></p>"
        updatephoto = "UPDATE TA_CORCP SET ta_descri = CASE WHEN RIGHT(TRIM(ta_descri), 4) = '<br>' THEN LEFT(TRIM(ta_descri), LENGTH(TRIM(ta_descri)) - 4) ELSE CONCAT(TRIM(ta_descri), '<br>') END, ta_images = '"+str(photo)+"' , ta_photo = '"+str(photo)+"' WHERE ta_cofou = '"+str(code_fou)+"' AND ta_socoda = '"+str(socoda)+"';"
        update = "UPDATE ta_corcp SET TA_DESCRI = '"+str(htmldesc)+"' WHERE ta_cofou = '"+str(code_fou)+"' AND ta_socoda = '"+str(socoda)+"';"
        if sheet2[colonne7 + str(tracer)].value != "REQ DESC":
            sheet2[colonne7 + str(tracer)].value = htmldesc
        if sheet2[colonne9 + str(tracer)].value != "REQ PHOTO":
            sheet2[colonne9 + str(tracer)].value = updatephoto
    
    
    outputfolder = wb.replace(".xlsx","_req.xlsx")
    workbook.save(outputfolder)
    end_time = time.perf_counter()
    execution_time = end_time - start_time
    print("----------------------------------------------------------")
    print("Temps d'exécution : {:.2f} seconds".format(execution_time))
    print("----------------------------------------------------------") 
    root.destroy()

def recuperer_ltre(valeur, plage):
    for row in plage:
        for cell in row:
            if cell.value == valeur :
                column = cell.column
                ltr_column = openpyxl.utils.get_column_letter(column)
                return ltr_column
        
def open_file():
    folder_path = filedialog.askopenfilename()
    folder_entry.delete(0, tk.END)
    folder_entry.insert(0, folder_path)

def update_progressbar(pbar, value):
    pbar['value'] = value
    pbar.update()

def clear_entry():
    folder_entry.delete(0, tk.END)
    valider_button.config(state='normal')
    folder_entry.config(state='normal')
    folder_button.config(state='normal')
    
def new_tarif():
    folder_entry.delete(0, tk.END)
    valider_button.config(state='normal')
    folder_entry.config(state='normal')
    folder_button.config(state='normal')
    clear_button.config(state='normal')

def OnValidation():
    folder_path = folder_entry.get()
    fourn_name = supplier_var.get()
    marq_name = brand_var.get()
    if folder_path != "":
        print("Chemin du dossier : "+folder_path)       
        if str(fourn_name) != "" or str(fourn_name) != "-Choisir un fournisseur-":
            print("Nom du fournisseur : "+fourn_name)
            if str(marq_name) != "" or str(marq_name) !="-Choisir une marque-":
                print("Nom de la marque : "+marq_name)
            else:
                print("Le nom de la marque n'est pas spécifié")
        else:
            print("Le nom du fournisseur n'est pas spécifié")
    else:
        print("Chemin du dossier n'est pas spécifié")
    
    if folder_path != "" and fourn_name != "" and marq_name != "":
        valider_button.config(state='disabled')
        folder_entry.config(state='disabled')
        folder_button.config(state='disabled')
        clear_button.config(state = "disabled")
        
        trigramme = get_trigram(fourn_name, marq_name).lower()
        create_request(folder_path, fourn_name, marq_name, trigramme)
         
def quit_app():
    root.destroy()
    
root = tk.Tk()
root.title("REQUETATOR23.07.11")
root.geometry("450x160")

texte_param = tk.Label(root, text="Initialisation de Requetator")
texte_param.grid(row=0, column=1)

folder_label = tk.Label(root, text="Fichier Tarif :")
folder_label.grid(row=1, column=0)

folder_entry = tk.Entry(root, width=30)
folder_entry.grid(row=1, column=1)

folder_button = tk.Button(root, text="   ...   ", command=open_file)
folder_button.grid(row=1, column=2)

supplier_label = tk.Label(root, text="Fournisseur")
supplier_label.grid(row=3, column=0)

supplier_var = tk.StringVar(root)
supplier_var.set("-Choisir un fournisseur-")

supplier_list = tk.OptionMenu(root, supplier_var, *df['FABRICANT'].unique())
supplier_list.grid(row=3, column=1)

brand_label = tk.Label(root, text="Marque")
brand_label.grid(row=4, column=0)

brand_var = tk.StringVar(root)
brand_var.set("-Choisir une marque-")

brand_list = tk.OptionMenu(root, brand_var, "")
brand_list.grid(row=4, column=1)

clear_button = tk.Button(root, text="Effacer", command=clear_entry, state= 'normal')
clear_button.grid(row=6, column=1)

valider_button = tk.Button(root, text="Valider",state='normal', command=OnValidation)
valider_button.grid(row=6, column=2)

quit_button = tk.Button(root, text="Quitter", command=quit_app)
quit_button.grid(row=7, column=2)

new_button = tk.Button(root, text="Nouveau ! ", command=new_tarif)
new_button.grid(row=7, column=0)

supplier_var.trace('w', update_brand_list)

root.mainloop()
