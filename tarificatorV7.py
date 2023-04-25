# -*- coding: utf-8 -*-
"""
Spyder Editor

Created on Tue Dec 13 13:43:25 2022

@author: cail
"""
import time
import os
import re
import string
import subprocess
import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows
import shutil
from win32com.client import Dispatch
import openpyxl
from openpyxl import load_workbook
import tkinter as tk
from tkinter import filedialog
from tkinter import messagebox

columns_gard = ["MARQUE", "REFCIALE", "REFARTICLE", "GTIN13", "LIBELLE30", "LIBELLE80",
                    "TARIF", "TARIFD", "QMV" ,"QMC" ,"QT" ,"UB" ,
                    "FAM1",	"FAM2", "FAM3", "MKT1", "MKT2", "MKT3", "LIBELLE240", "STA"]

columns_gard_deee = ["REFCIALE", "RNBR", "RVAL", "RNUM", "RCOD"]

columns_gard_socoda = ["REFCIALE", "SKUSOCODA"]

columns_gard_media = ['REFCIALE', 'NUM', 'NOM']

columns_gard_fgaz = ['REFCIALE', 'RCOD']

mois = ['01', '02', '03', '04', '05', '06', '07', '08', '09', '10', '11', '12']

annees = ['2023', '2024', '2025', '2026']

FINAL = ""

df = pd.read_excel('PrefixeSocoda (003).xlsx')
supplier_to_brands = {}
for supplier, brand in zip(df['FABRICANT'], df['MARQUE']):
    if supplier not in supplier_to_brands:
        supplier_to_brands[supplier] = []
    if brand not in supplier_to_brands[supplier]:
        supplier_to_brands[supplier].append(brand)

# Mettre à jour la liste des marques en fonction du fournisseur sélectionné
def update_brand_list(*args):
    selected_supplier = supplier_var.get()
    brand_list['menu'].delete(0, 'end')
    brands = df[df['FABRICANT'] == selected_supplier]['MARQUE'].unique()
    for brand in brands:
        brand_list['menu'].add_command(label=brand, command=tk._setit(brand_var, brand))

def get_trigram(fabricant, marque):
    trigram = df.loc[(df['FABRICANT'] == fabricant) & (df['MARQUE'] == marque), 'PREFIXE'].values
    if len(trigram) == 0:
        return "Trigramme non trouvé"
    else:
        return trigram[0]

def get_compatible(fabricant, marque):
    comp = df.loc[(df['FABRICANT'] == fabricant) & (df['MARQUE'] == marque), 'COMPATIBLE'].values
    if len(comp) == 0:
        return "Compatibilité non trouvé"
    else:
        return comp[0]
    
def show_error_popup(message):
    root = tk.Tk()
    root.withdraw()
    messagebox.showerror("Error", message)
    
def keep_only_numbers(string):
  return ''.join(char for char in string if char.isdigit())

def remove_numbers(string):
    return ''.join(c for c in string if c.isalpha())

def recup_media_name(worksheet_names):
    for media in worksheet_names:
        if media.startswith('MEDIA'):
            return media

def recup_sheet_index(string, file):
    workbook = openpyxl.load_workbook(file)
    worksheet_names = workbook.sheetnames
    for sheet in worksheet_names:
        if string in sheet:
            index = worksheet_names.index(sheet)
            return index


# Barre de progression

def printProgressBar (iteration, total, prefix = '', suffix = '', decimals = 1, length = 100, fill = '█', printEnd = "\r"):

    percent = ("{0:." + str(decimals) + "f}").format(100 * (iteration / float(total)))
    filledLength = int(length * iteration // total)
    bar = fill * filledLength + '-' * (length - filledLength)

    print(f'\r{prefix} |{bar}| {percent}% {suffix}', end = printEnd)
    # Print New Line on Complete
    if iteration == total:
        print()

# Permet de récuperer la lettre d'une colonne à partir de
# la valeur de la colonne et la plage de recherche

def recuperer_ltre(valeur, plage):
    for row in plage:
        for cell in row:
            if cell.value == valeur :
                column = cell.column
                ltr_column = openpyxl.utils.get_column_letter(column)
                return ltr_column

# Permet de faire une recherche Vertical pour remplir les colonnes
# PHOTO, FICHE, SKUSOCODA

def camsoule(filename, col_name, onglet, destfile):
    print("----------------------------------------------------------")
    print('Recherche en cours : '+col_name)
    print("----------------------------------------------------------")
    wb = openpyxl.load_workbook(filename)
    sheet = wb['01_COMMERCE']
    sheet2 = wb[onglet]
    plage = str(sheet2.dimensions)
    plage = remove_numbers(plage)
    plage = plage[:1]+':'+plage[1:]
    plage = sheet2[plage]
    lettre = recuperer_ltre(col_name, plage)
    
    ltr_fgaz = recuperer_ltre('F-GAZ', sheet)
    ltr_deee = recuperer_ltre('D3E', sheet)
    ltr_d3ev = recuperer_ltre("D3EV", sheet)
    ltr_d3eu = recuperer_ltre("D3EU", sheet)
    ltr_d3ec = recuperer_ltre("D3EC", sheet)
    ltr_photo = recuperer_ltre("PHOTO", sheet)
    ltr_skusocoda = recuperer_ltre("SKUSOCODA", sheet)
    ltr_fiche = recuperer_ltre("FICHE", sheet)
    lettre_refartcl = recuperer_ltre('REFCIALE', sheet)

    if lettre_refartcl == None:
        raise ValueError("La colonne : REFCIALE, n'existe pas. ")
    colonne = sheet[str(lettre_refartcl)]
    max_row = sheet.max_row
    printProgressBar(0, max_row, prefix = 'Progress:', suffix = 'Complete', length = 50)
    for row in colonne:
        valuer = row.value
        tracer = row.row

        printProgressBar(tracer, max_row, prefix = 'Progress:', suffix = 'Complete', length = 50)
        for ligne in plage:
            for cell in ligne:
                    if str(cell.value) == str(valuer):
                        ligne = cell.row
                        resultat = sheet2[lettre+str(ligne)].value
                        if col_name == 'PHOTO':
                            sheet[ltr_photo+str(tracer)].value = resultat
                        if col_name == 'FICHE':
                            sheet[ltr_fiche+str(tracer)].value = resultat
                        if col_name == 'SKUSOCODA':
                            sheet[ltr_skusocoda+str(tracer)].value = resultat
                        if col_name == 'D3EC':
                            if sheet[ltr_deee+str(tracer)].value == None:
                                sheet[ltr_deee+str(tracer)].value = "*"
                            sheet[ltr_d3ec+str(tracer)].value = resultat
                        if col_name == 'D3EU':
                            sheet[ltr_d3eu+str(tracer)].value = resultat
                        if col_name == 'D3EV':
                            sheet[ltr_d3ev+str(tracer)].value = resultat
                        if col_name == 'RCOD':
                            sheet[ltr_fgaz+str(tracer)].value = resultat

    wb.save(destfile)

# Copie colle le dossier original pour ne pas corrompre le dossier original
# Récupere par la meme occasion le FAB-DIS

def parse_folder(working_folder, columns_supr, four_name, tarif_date, trigramme, file_path, origin_folder=None):
    start_time = time.perf_counter()
    if origin_folder:
        if not os.path.exists(origin_folder):
            show_error_popup("Le dossier d'origine n'existe pas. ")
            raise ValueError("Le dossier d'origine n'existe pas. ")
        if os.path.exists(working_folder):
            shutil.rmtree(working_folder)
        print("----------------------------------------------------------")
        print("Copie en cours ... ")
        print("----------------------------------------------------------")

        shutil.copytree(origin_folder, working_folder)

    folder_content = os.listdir(working_folder)
    fichier_fabdis = file_path
    fichier_skusocoda = ""
    for element in folder_content:
        minus = element.lower()
        if minus.startswith("socoda"):
            fichier_skusocoda = os.path.join(working_folder, element)
    if fichier_skusocoda == None or fichier_skusocoda == "":
        show_error_popup("Le fichier socoda est introuvable. ")
        raise ValueError("Le fichier socoda est introuvable. ")

    log_file = os.path.join(working_folder, "logfile.txt")
    log_file = open(log_file, "w")

    destfile = os.path.join(working_folder, "Tarif_"+four_name+"_"+tarif_date+"_travail.xlsx")
    workbook = openpyxl.Workbook()
    workbook.save(destfile)

    log_file.write("FICHIER LOG TARIFICATOR \n\n\n")
    log_file.write("Commencement du script sur le fournisseur : " +four_name+ "\n")
    create_work_file(fichier_fabdis, columns_supr, four_name, destfile, log_file, trigramme, fichier_skusocoda, start_time)


# Création de mon fichier "Travail" qui copie colle mon Onglet 01-COMMERCE
def create_work_file(fabdis_file, columns_supr, four_name, destfile, log_file, trigramme, fichier_skusocoda, start_time):
    print("----------------------------------------------------------")
    print("Fichier FAB-DIS : "+fabdis_file)
    print("----------------------------------------------------------")
    print("----------------------------------------------------------")
    print("Ajustement des colonnes ...")
    print("----------------------------------------------------------")
    dfs = pd.read_excel(fabdis_file, sheet_name=None)

    df_commerce = dfs["01_COMMERCE"]
    df_commerce = df_commerce.loc[:, columns_gard]
    df_commerce = df_commerce.loc[~df_commerce['STA'].str.startswith('S')]
    if four_name != "HIKVISION FRANCE":
        df_commerce["FAM1"] = pd.to_numeric(df_commerce["FAM1"], errors="coerce")
        df_commerce["FAM1"] = df_commerce["FAM1"].apply(lambda x: '{:03d}'.format(x) if not pd.isna(x) else '')
    df_commerce["REFCIALE"] = df_commerce["REFCIALE"].astype(str)
    df_commerce["REFCIALE"] = df_commerce["REFCIALE"].str.zfill(6)
    df_commerce["REFARTICLE"] = df_commerce["REFARTICLE"].astype(str)
    df_commerce["REFARTICLE"] = df_commerce["REFARTICLE"].str.zfill(6)
    
    if four_name == "ATLANTIC CLIMATISATION & VENTILATION":
        df_commerce = df_commerce.loc[~df_commerce['MKT1'].str.startswith('CV4')]
    
    workbook = openpyxl.load_workbook(destfile)
    new_sheet = workbook.create_sheet("01_COMMERCE")
    for r in dataframe_to_rows(df_commerce, index=False, header=True):
        new_sheet.append(r)
        
    workbook.save(destfile)
    
    max_row, max_col = df_commerce.shape
    max_col_letters = string.ascii_uppercase[max_col-1]
    rep  = str(max_col_letters) + ":"+ str(max_row) 
    rep = "A:" + remove_numbers(rep)
    format_work_file(destfile, columns_supr, log_file, fichier_skusocoda, trigramme, four_name, start_time, fabdis_file, rep)

# Suppression des colonnes inutile puis insertion des colonnes
# PHOTO, FICHE, SKUSOCODA
# Copie du SKUSOCODA dans l'onglet SKUSOCODA
# Copie de tout l'onglet MEDIA
def format_work_file(destfile, columns_gard, log_file, fichier_skusocoda, trigramme, four_name, start_time, fabdis_file, rep):
    SKUSOCODA = False
    nom = 'NOM'
    #-------------------------------------------TRAVAUX-------------------------------------
    print("Mise en place de l'onglet D3E ")
    dfs = pd.read_excel(fabdis_file, sheet_name=None)

    df_deee = dfs["04_REGLEMENTAIRE"]
    df_deee = df_deee[df_deee["RTYP"] == 'CONTRIB']
    df_deee = df_deee.loc[:, columns_gard_deee]
    df_deee["REFCIALE"] = df_deee["REFCIALE"].astype(str)
    df_deee["REFCIALE"] = df_deee["REFCIALE"].str.zfill(6)
    try:
        df_deee.loc[~df_deee['RCOD'].str.startswith('P')]
    except Exception :
        print("")
    else:
        df_deee = df_deee.loc[~df_deee['RCOD'].str.startswith('P')]

    workbook = openpyxl.load_workbook(destfile)
    new_sheet = workbook.create_sheet("DEEE")
    for r in dataframe_to_rows(df_deee, index=False, header=True):
        new_sheet.append(r)
    workbook.save(destfile)
    
    #-------------------------------------------TRAVAUX-------------------------------------
    print("Mise en place de l'onglet MEDIA  ")
    workbook = load_workbook(fabdis_file)
    sheet4 = workbook['03_MEDIA']
    photobd = ['photobd', 'PHOTOBD']
    photohd = ['photohd', 'PHOTOHD']
    photohda = ['photohda', 'PHOTOHDA']
    photonorm = ['photo', 'PHOTO']
    photoaprendre = ''
    status = False
    colonne2 = recuperer_ltre('TYPM', sheet4['A:Z'])
    for row in sheet4[colonne2]:
        if row.value == photobd:
            photoaprendre = row.value
            status = True
            break
        
    if status == False: 
        for row in sheet4[colonne2]:
            if row.value in photohd:
                photoaprendre = row.value
                status = True
                break
    
    if status == False: 
        for row in sheet4[colonne2]:
            if row.value in photohda:
                photoaprendre = row.value
                status = True
                break
            
    if status == False: 
        for row in sheet4[colonne2]:
            if row.value in photonorm:
                photoaprendre = row.value
                status = True
                break
    print(photoaprendre)
    dfs = pd.read_excel(fabdis_file, sheet_name=None)
    df_media = dfs["03_MEDIA"]
    df_media = df_media[df_media["TYPM"] == photoaprendre]
    df_media = df_media.loc[:, columns_gard_media]
    df_media = df_media[df_media["NUM"] == 1]
    df_media["REFCIALE"] = df_media["REFCIALE"].astype(str)
    df_media["REFCIALE"] = df_media["REFCIALE"].str.zfill(6)

    workbook = openpyxl.load_workbook(destfile)
    new_sheet = workbook.create_sheet("03_MEDIA")
    for r in dataframe_to_rows(df_media, index=False, header=True):
        new_sheet.append(r)
    workbook.save(destfile)

    #-------------------------------------------TRAVAUX-------------------------------------
    print("Mise en place de l'onglet F-GAZ")
    dfs = pd.read_excel(fabdis_file, sheet_name=None)
    df_fgaz = dfs["04_REGLEMENTAIRE"]
    df_fgaz = df_fgaz[df_fgaz["RTYP"] == "F-GAZ"]
    df_fgaz = df_fgaz[df_fgaz["RTEXTE"] == "SOUMIS ADC"]
    df_fgaz = df_fgaz.loc[:, columns_gard_fgaz]
    df_fgaz["REFCIALE"] = df_fgaz["REFCIALE"].astype(str)
    df_fgaz["REFCIALE"] = df_fgaz["REFCIALE"].str.zfill(6)

    workbook = openpyxl.load_workbook(destfile)
    new_sheet = workbook.create_sheet("F-GAZ")
    for r in dataframe_to_rows(df_fgaz, index=False, header=True):
        new_sheet.append(r)

    workbook.save(destfile)

    #-------------------------------------------TRAVAUX-------------------------------------
    workbook = load_workbook(destfile)

    sheet3 = workbook['03_MEDIA']
    sheet = workbook['01_COMMERCE']
    sheet2 = workbook['DEEE']

    print("Création des bons noms d'images ")

    col_fiche = 4
    sheet3.cell(row = 1, column = col_fiche, value="FICHE")
    ext2 = '.pdf'
    column6 = recuperer_ltre(nom, sheet3['A:T'])
    column7 = recuperer_ltre('REFCIALE', sheet3['A:T'])
    column8 = recuperer_ltre('FICHE', sheet3['A:T'])
    for row in sheet3[column6]:
        tracer = row.row
        if row.value != nom:
            ext = ".jpg"
            ref = sheet3[column7 + str(tracer)].value
            photo_new_name = trigramme +'_'+ str(ref) + ext
            fiche_new_name = trigramme +'_'+ str(ref) + ext2
            row.value = photo_new_name
            sheet3[column8 + str(tracer)].value = fiche_new_name
        if row.value == nom:
            row.value = 'PHOTO'

    print("Etape 9")

    column9 = recuperer_ltre('LIBELLE240', sheet[rep])
    column10 = recuperer_ltre('LIBELLE30', sheet[rep])
    for row in sheet[column9]:
        rowname = str(row.value)
        rowname = rowname.replace("œ","oe")
        row.value = rowname
    for row in sheet[column10]:
        rowname = str(row.value)
        rowname = rowname.replace("œ","oe")
        row.value = rowname

    print("----------------------------------------------------------")
    print("Insertion des colonnes 'PHOTO', 'FICHE', D3E et 'SKUSOCODA' ...")
    print("----------------------------------------------------------")
    
    
    sheet.insert_cols(idx = sheet.max_column+1, amount=11)
    
    col_socoda = sheet.max_column+1
    sheet.cell(row = 1, column = col_socoda, value="SOCODA")
    col_uch = sheet.max_column+1
    sheet.cell(row = 1, column = col_uch, value="UCH")
    col_fgaz = sheet.max_column+1
    sheet.cell(row = 1, column = col_fgaz, value="F-GAZ")
    col_photo = sheet.max_column+1
    sheet.cell(row = 1, column = col_photo, value="PHOTO")
    col_fiche = sheet.max_column+1
    sheet.cell(row = 1, column = col_fiche, value="FICHE")
    col_skusocoda = sheet.max_column+1
    sheet.cell(row = 1, column = col_skusocoda, value="SKUSOCODA")
    col_d3e = sheet.max_column+1
    sheet.cell(row = 1, column = col_d3e, value="D3E")
    col_d3ec = sheet.max_column+1
    sheet.cell(row = 1, column = col_d3ec, value="D3EC")
    col_d3ev = sheet.max_column+1
    sheet.cell(row = 1, column = col_d3ev, value="D3EV")
    col_d3eu = sheet.max_column+1
    sheet.cell(row = 1, column = col_d3eu, value="D3EU")
    col_deee = sheet2.max_column+1
    sheet2.cell(row = 1, column = col_deee, value="D3EC")

    log_file.write("Les colonnes suivantes ont été inséré : 'PHOTO', 'FICHE', D3E, UCH, D3EC, D3EV, D3EU et 'SKUSOCODA'  \n ")

    
    print("Etape 11")
    column12 = recuperer_ltre('SOCODA', sheet['A:AZ'])
    column13 = recuperer_ltre('REFARTICLE', sheet['A:AZ'])
    for row in sheet[column13]:
        tracer = row.row
        if row.value != "" and str(row.value) != 'REFARTICLE':
            sheet[column12 + str(tracer)].value = str(trigramme) + str(row.value)


    #Mise en place de la D3E

    ltr_deee = recuperer_ltre('D3EC', sheet2['A1:F1'])
    ltr_rcod = recuperer_ltre('RCOD', sheet2['A1:F1'])
    ltr_rnbr = recuperer_ltre('RNBR', sheet2['A1:F1'])
    ltr_rval = recuperer_ltre('RVAL', sheet2['A1:F1'])
    ltr_rnum = recuperer_ltre('RNUM', sheet2['A1:F1'])
    ltr_ref = recuperer_ltre('REFCIALE', sheet2['A1:F1'])

    colonne = sheet2[str(ltr_rcod)]
    colonne2 = sheet2['A1:F1']
    colonne3 = sheet2[str(ltr_rnum)]
    colonne4 = sheet2[str(ltr_ref)]
    val_deee_1 = ""
    val_deee_2 = ""

    # Detection de DEEE multiple et addition des valeurs
    for val in colonne3:
        tracer = val.row
        if str(val.value) == '2':
            val_deee_2 = sheet2[ltr_rnbr+str(tracer)].value * float(sheet2[ltr_rval+str(tracer)].value)
            get_ref = sheet2[ltr_ref+str(tracer)].value
            for ref in colonne4:
                tracer2 = ref.row
                if str(ref.value) == str(get_ref) and str(sheet2[ltr_rnum+str(tracer2)].value) == "1" :
                    val_deee_1 = float(sheet2[ltr_rval+str(tracer2)].value) * sheet2[ltr_rnbr+str(tracer2)].value
                    sheet2[ltr_deee+str(tracer2)].value = "ECO_"+ str(int((val_deee_2 + val_deee_1)*100))

    # Suppression des lignes de valeur 2

    for val in colonne3:
        tracer = val.row
        if str(val.value) == '2':
            sheet2.delete_rows(idx = tracer, amount=1)

    for headers in colonne2:
        for header in headers:
            if sheet2[recuperer_ltre(header.value, sheet2['A:F'])+"1"].value == 'RNBR' :
                sheet2[recuperer_ltre(header.value, sheet2['A:F'])+"1"].value = 'D3EU'
            if sheet2[recuperer_ltre(header.value, sheet2['A:F'])+"1"].value == 'RVAL' :
                sheet2[recuperer_ltre(header.value, sheet2['A:F'])+"1"].value = 'D3EV'
    for row in colonne:
        tracer = row.row
        deee = sheet2[ltr_deee+str(tracer)].value
        rcod = sheet2[ltr_rcod+str(tracer)].value
        rnbr = sheet2[ltr_rnbr+str(tracer)].value
        rval = sheet2[ltr_rval+str(tracer)].value
        rval = str(rval).replace(",", ".") 
        if deee == None:
            if str(rcod).startswith("L") and float(rval) > 0.01:
                if float(rval) == 0.13:
                    if int(rnbr) < 10 :
                        sheet2[ltr_deee+str(tracer)].value = "ECL0" + str(rnbr)
                    else:
                        sheet2[ltr_deee+str(tracer)].value = "ECL" + str(rnbr)

                if float(rval) == 0.10:
                    if int(rnbr) < 10 :
                        sheet2[ltr_deee+str(tracer)].value = "ECL0"+str(rnbr)+"R"
                    else:
                        sheet2[ltr_deee+str(tracer)].value = "ECL" +str(rnbr)+"R"
            else:
                sheet2[ltr_deee+str(tracer)].value = "ECO_"+str(int(int(rnbr)*float(rval)*100))

    log_file.write("La D3E à bien été pris en charge \n ")
    # Mise en place de UCH

    ltr_uch = recuperer_ltre('UCH', sheet['A1:AZ1'])
    ltr_qt = recuperer_ltre('QT', sheet['A1:AZ1'])
    ltr_ub = recuperer_ltre('UB', sheet['A1:AZ1'])

    colonne = sheet[str(ltr_qt)]
    nbr_except = 0
    for cell in colonne:
        tracer = cell.row
        if str(cell.value) == '1':
            if sheet[ltr_ub+str(tracer)].value == 'PF':
                sheet[ltr_uch+str(tracer)].value = 'PI'
            if sheet[ltr_ub+str(tracer)].value == 'EA':
                sheet[ltr_uch+str(tracer)].value = 'PI'
            if sheet[ltr_ub+str(tracer)].value == 'LTR':
                sheet[ltr_uch+str(tracer)].value = 'PI'
            if sheet[ltr_ub+str(tracer)].value == 'MTR':
                sheet[ltr_uch+str(tracer)].value = 'ME'
            if sheet[ltr_ub+str(tracer)].value == 'PK':
                sheet[ltr_uch+str(tracer)].value = 'PI'

            if sheet[ltr_ub+str(tracer)].value != 'EA' and sheet[ltr_ub+str(tracer)].value != 'LTR' and sheet[ltr_ub+str(tracer)].value != 'MTR' and sheet[ltr_ub+str(tracer)].value != 'PK' and sheet[ltr_ub+str(tracer)].value != 'PF' :
                sheet[ltr_uch+str(tracer)].value = 'Exept'
                nbr_except = nbr_except +1

        if str(cell.value) == '100':
            if sheet[ltr_ub+str(tracer)].value == 'PF':
                sheet[ltr_uch+str(tracer)].value = 'PI'
            if sheet[ltr_ub+str(tracer)].value == 'BX':
                sheet[ltr_uch+str(tracer)].value = 'CP'
            if sheet[ltr_ub+str(tracer)].value == 'EA':
                sheet[ltr_uch+str(tracer)].value = 'CP'
            if sheet[ltr_ub+str(tracer)].value == 'MTR':
                sheet[ltr_uch+str(tracer)].value = 'CM'
            if sheet[ltr_ub+str(tracer)].value == 'PK':
                sheet[ltr_uch+str(tracer)].value = 'CP'
            if sheet[ltr_ub+str(tracer)].value == 'SA':
                sheet[ltr_uch+str(tracer)].value = 'CP'

            if sheet[ltr_ub+str(tracer)].value != 'BX' and sheet[ltr_ub+str(tracer)].value != 'EA' and sheet[ltr_ub+str(tracer)].value != 'MTR' and sheet[ltr_ub+str(tracer)].value != 'PK' and sheet[ltr_ub+str(tracer)].value != 'SA' and sheet[ltr_ub+str(tracer)].value != 'PF':
                print(sheet[ltr_ub+str(tracer)].value)
                sheet[ltr_uch+str(tracer)].value = 'Exept'
                nbr_except = nbr_except +1

        if str(cell.value) == '1000':
            if sheet[ltr_ub+str(tracer)].value == 'EA':
                sheet[ltr_uch+str(tracer)].value = 'MI'

            if sheet[ltr_ub+str(tracer)].value != 'EA':
                sheet[ltr_uch+str(tracer)].value = 'Exept'
                nbr_except = nbr_except +1

    log_file.write("L'indice UCH à été mis à jour \n ")
    log_file.write("Le nombre d'exception UCH est de : "+str(nbr_except)+ "\n ")
    # VERIFICATOR


    workbook.save(destfile)
    print("----------------------------------------------------------")
    print("Copie SKUSOCODA ...")
    print("----------------------------------------------------------")

    try : 
        pd.read_excel(fichier_skusocoda, sheet_name=None)
    except Exception :
        print("")
    else:
        dfs = pd.read_excel(fichier_skusocoda, sheet_name=None)
        df_skusocoda = dfs['S1_SOCODA_NOMENCLATURE']
        df_skusocoda = df_skusocoda.loc[:, ["REFCIALE", "SKUSOCODA"]]
        df_skusocoda["REFCIALE"] = df_skusocoda["REFCIALE"].astype(str)
        df_skusocoda["REFCIALE"] = df_skusocoda["REFCIALE"].str.zfill(6)
        SKUSOCODA = True
        workbook = openpyxl.load_workbook(destfile)
        new_sheet = workbook.create_sheet("SKUSOCODA")
        for r in dataframe_to_rows(df_skusocoda, index=False, header=True):
            new_sheet.append(r)
        workbook.save(destfile)
    
    
    camsoule(destfile, "RCOD", 'F-GAZ', destfile)
    log_file.write("Recherche V des F-GAZ terminé ! \n ")
    
    # parcourir chaque ref pour verifier si elle est dans le media
    if SKUSOCODA == True:
        camsoule(destfile, "SKUSOCODA", 'SKUSOCODA', destfile)
        log_file.write("Recherche V des SKUSOCODA terminé ! \n ")
    
    
    camsoule(destfile, "PHOTO", '03_MEDIA', destfile)
    log_file.write("Recherche V des PHOTO terminé ! \n ")
    camsoule(destfile, "FICHE", '03_MEDIA', destfile)
    log_file.write("Recherche V des FICHE terminé ! \n ")
    camsoule(destfile, "D3EC", 'DEEE', destfile)
    log_file.write("Recherche V des DEEE terminé ! \n ")
    camsoule(destfile, "D3EV", 'DEEE', destfile)
    log_file.write("Recherche V des D3EV terminé ! \n ")
    camsoule(destfile, "D3EU", 'DEEE', destfile)
    log_file.write("Recherche V des D3EU terminé ! \n ")
    log_file.write("______________________________________________________________________________ \n ")
    log_file.close()

    tarif_date = annee_selectionnee.get() + mois_selectionne.get()
    FINAL = os.path.join(os.path.dirname(destfile), "Tarif_"+four_name+"_"+tarif_date+".xlsx")
    workbook = openpyxl.Workbook()
    workbook.save(FINAL)
    index_commerce = recup_sheet_index('01_COMMERCE', destfile)+1

    xl = Dispatch("Excel.Application")
    wb1 = xl.Workbooks.Open(Filename=destfile)
    wb2 = xl.Workbooks.Open(Filename=FINAL)
    ws1 = wb1.Worksheets(index_commerce)
    ws1.Copy(Before=wb2.Worksheets(1))
    wb2.Close(SaveChanges=True)
    xl.Quit()
    
    end_time = time.perf_counter()
    execution_time = end_time - start_time
    print("----------------------------------------------------------")
    print("Temps d'exécution : {:.2f} seconds".format(execution_time))
    print("----------------------------------------------------------")
    start_popup(FINAL)
if __name__ == '__main__':
    # input
    def open_folder():
        file_path = filedialog.askopenfilename()
        folder_entry.delete(0, tk.END)
        folder_entry.insert(0, file_path)

    def clear_entry():
        folder_entry.delete(0, tk.END)
        supplier_var.set("-Choisir un fournisseur-")
        brand_var.set("-Choisir une marque-")
        valider_button.config(state='normal')
        folder_entry.config(state='normal')
        folder_button.config(state='normal')
        supplier_list.config(state='normal')

    def start_mediator():
        subprocess.call(["python", "mediatorV2.py"])
        
    def start_popup(FINAL):
        print("Le fichier final est : " + FINAL)
        subprocess.call(["python", "popup.py", FINAL])
        
    def OnValidation():
        file_path = folder_entry.get()
        fourn_name = supplier_var.get()
        marq_name = brand_var.get()
        tarif_date = annee_selectionnee.get() + mois_selectionne.get()
        if file_path != "":
            print("Chemin du dossier : "+file_path)
            if str(fourn_name) != "" or str(fourn_name) != "-Choisir un fournisseur-":
                print("Nom du fournisseur : "+fourn_name)
                if str(marq_name) != "" or str(marq_name) !="-Choisir une marque-":
                    print("Nom de la marque : "+marq_name)
                else:
                    print("Le nom de la marque n'est pas spécifié")
            else:
                print("Le nom du fournisseur n'est pas spécifié")
        else:
            print("Chemin du dossier n'est pas spécifié")

        if file_path != "" and fourn_name != "":
             valider_button.config(state='disabled')
             folder_entry.config(state='disabled')
             folder_button.config(state='disabled')
             supplier_list.config(state='disabled')
             clear_button.config(state = "disabled")
             mediator_bouton.config(state= "disabled")
             mois_menu.config(state= 'disabled')
             annee_menu.config(state= 'disabled')
             brand_list.config(state= 'disabled')
             folder_path = os.path.dirname(file_path)
             output_folder = folder_path+"-2"

             trigramme = get_trigram(fourn_name, marq_name)
             comp = get_compatible(fourn_name, marq_name)
             if comp == "NON":
                 show_error_popup("Ce tarif n'est pas compatible avec le logiciel !")
             if comp == "OUI":
                 parse_folder(output_folder, columns_gard, fourn_name, tarif_date, trigramme, file_path, folder_path)
             
    def new_tarif():
        folder_entry.delete(0, tk.END)
        valider_button.config(state='normal')
        folder_entry.config(state='normal')
        folder_button.config(state='normal')
        supplier_list.config(state='normal')
        clear_button.config(state='normal')
        mois_menu.config(state= 'normal')
        annee_menu.config(state= 'normal')
        brand_list.config(state= 'normal')
        mediator_bouton.config(state= 'normal')

    def quit_app():
        root.destroy()

    root = tk.Tk()
    root.title("TARIFICATOR 7.8.0")
    root.geometry("560x220")

    mois_selectionne = tk.StringVar(root)
    annee_selectionnee = tk.StringVar(root)
    mois_selectionne.set(mois[0])
    annee_selectionnee.set(str(annees[0]))

    texte_param = tk.Label(root, text="Initialisation de Tarificator")
    texte_param.grid(row=0, column=1)

    folder_label = tk.Label(root, text="Dossier contenant le FAB-DIS :")
    folder_label.grid(row=1, column=0)

    folder_entry = tk.Entry(root, width=30)
    folder_entry.grid(row=1, column=1)

    folder_button = tk.Button(root, text="   ...   ", command=open_folder)
    folder_button.grid(row=1, column=2)

    supplier_label = tk.Label(root, text="Fournisseur")
    supplier_label.grid(row=3, column=0)

    supplier_var = tk.StringVar(root)
    supplier_var.set("-Choisir un fournisseur-")

    supplier_list = tk.OptionMenu(root, supplier_var, *df['FABRICANT'].unique())
    supplier_list.grid(row=3, column=1)

    brand_label = tk.Label(root, text="Marque")
    brand_label.grid(row=4, column=0)

    brand_var = tk.StringVar(root)
    brand_var.set("-Choisir une marque-")

    brand_list = tk.OptionMenu(root, brand_var, "")
    brand_list.grid(row=4, column=1)

    date_label = tk.Label(root, text="Date du tarif :")
    date_label.grid(row=5, column=0)

    mois_label = tk.Label(root, text="Mois :")
    mois_label.grid(row=5, column=0,sticky=tk.E)

    mois_menu = tk.OptionMenu(root, mois_selectionne, *mois)
    mois_menu.grid(row=5, column=1, sticky=tk.W)

    annee_label = tk.Label(root, text="Année :")
    annee_label.grid(row=5, column=1)

    annee_menu = tk.OptionMenu(root, annee_selectionnee, *annees)
    annee_menu.grid(row=5, column=1, sticky=tk.E)

    expt_label = tk.Label(root, text="Exporter les images et fiche du FAB-DIS :")
    expt_label.grid(row=6, column=0)

    mediator_bouton = tk.Button(root, text="Mediator", command=start_mediator, state= 'normal')
    mediator_bouton.grid(row=6, column=1, sticky=tk.W, padx= 20)

    clear_button = tk.Button(root, text="Effacer", command=clear_entry, state= 'normal')
    clear_button.grid(row=7, column=1)

    valider_button = tk.Button(root, text="Valider",state='normal', command=OnValidation)
    valider_button.grid(row=7, column=2)

    quit_button = tk.Button(root, text="Quitter", command=quit_app)
    quit_button.grid(row=9, column=2)

    new_button = tk.Button(root, text="Nouveau Tarif ! ", command=new_tarif)
    new_button.grid(row=9, column=0)

    supplier_var.trace('w', update_brand_list)

    root.mainloop()
